# -*- coding: UTF-8 -*-
import re
import os
import sys 
import time
#import xlrd
#import xlwt
from openpyxl import Workbook
import platform

'''
	Read the csv file to compare with raw data
'''
def readCSV(csvFile):
	print 'Start read CSV...'
	dir = sys.path[0]
	dir_2 = dir.split('\\')
	fname = ""
	for i in range(len(dir_2)):
		fname += dir_2[i] + "\\\\"
	fnames = fname
	fnames += csvFile
	sysstr = platform.system()
	if(sysstr != "Windows"):
		fnames = csvFile
	print 'scvFile name ', csvFile 
	fp = open(fnames, 'r')
	res = []
	for line in fp:
		if line.find("Time;appname") is -1:
			l = line.split(';') 
			s1 = l[1].split(':')
			s1 = s1[0]
			s2 = l[2]
			sl = s2.split(' ')
			#url = s1 + sl[1] + ";" + sl[0]
			url = s1 + sl[1]
			url += ";" + sl[0]
			#print url
			res.append(url)
	fp.close()
	print 'Complete reading csv...'
	return res
'''
	Check the url in raw data is in the csv file too
'''
def isInCSV(url, matrix):
	return (url in matrix)

'''
	Create interfile to shorten the time of the program
'''
def CreateInterFile(delta, InputFile):
	print 'Start Create inter files...'
	dir = sys.path[0]
	dir_2 = dir.split('\\')
	fname = "" 
	str1 = "inter_delta"
	for i in range(len(dir_2)):
		fname += dir_2[i] + "\\\\" 
	fnames = fname
	fnames2 = fname
	fnames += InputFile
	fnames2 += str1
	sysstr = platform.system()
	if(sysstr != "Windows"):
		fnames = InputFile
		fnames2 = str1
		pass
	print "fnames", fnames
	fp = open(fnames, 'r') #Open the file of raw data
	fp2 = open(fnames2, 'w')
	#fp2.close()
	#fp2 = open(fnames2, 'a') 
	for line in fp:
		if (line.find("mybluemix.net") is -1):
			fp2.writelines(line)
	fp.close()
	fp2.close()
	print 'Create inter file done...'
	
'''
	the function to filter the raw data
	delta the unit of time , the distance between the timestamp
	if delta is 1800, it means 1800 seconds
	InputFile means the name of file that contains the raw data
'''
def solve(delta, InputFile, csvFile):
	insteadString = "api.ng.bluemix.net/v2/app"
	specString = "api.ng.bluemix.net/v2/apps/uuid_replaced;PUT"
	print 'delta is %d' %(delta)
	interFIleName = "inter_delta"
	dir = sys.path[0]
	dir_2 = dir.split('\\')
	fname = ""
	for i in range(len(dir_2)):
		fname += dir_2[i] + "\\\\"
	fnames = fname
	fnames += csvFile
	sysstr = platform.system()
	print "platform:",sysstr
	if(sysstr != "Windows"):
		fnames = csvFile
	if(sysstr == "Windows"):
		interFIleName = fname + interFIleName
	
	if(os.path.exists(interFIleName) is False):  #If interFIleName doesn't exist, then create it
		print 'Does not exit interFIle...'
		CreateInterFile(delta, InputFile)
	else:
		print 'InterFile exits...'
	print 'interfilename', interFIleName
	fp = open(interFIleName, 'r')
	
	dictNum = {} #Record the appearance of the every api
	dict = {} #Record the response time and latency which is saved as a tuple of all service

	earlyTime = 9999999999 
	latestTime = 0 
	
	patternStr = r'%s(.+?)%s' %('\"','\"')
	APIPatten = re.compile(patternStr, re.IGNORECASE) #Get the pattern to find the url of apis
	patternStr = r'%s(.+?)%s' %('response_time:', ' ')
	ResponseTimePattern = re.compile(patternStr, re.IGNORECASE) #Get the pattern to find the ResponseTime Pattern
	patternStr = r'%s(.+?)%s' %(' - ', ']')
	TimeStampPattern = re.compile(patternStr, re.IGNORECASE) #Get the pattern to find the ResponseTime Pattern

	csvMatrix = readCSV(csvFile) #Get matrix  list from csv file
	
	idx = 0
	for line in fp:
		#print idx
		time_st_1 = ""
		res_time = 0
		hasApi = False
		api = ""
		idx += 1
		formaterror = False
		if line.find("mybluemix.net") is not -1:
			continue
		#######
		try:
			(i, j) = re.search(' - ', line).span()
			api = line[0: i] 
			post = re.search(APIPatten, line) 
			tmpPost = post.group(0)
			tl = len(tmpPost) - 1
			tmpPost = tmpPost[1:tl] 
			list = tmpPost.split(" ")
			#api += list[1] + ";" + list[0] #get the full name of url
			
			api += list[1] 
			api = api.split('?')
			api = api[0] 
			api += ";" + list[0]
			bk = False
			if (api.find(insteadString) is not -1): 
				api = insteadString	+ ";PUT"
				bk = True
				
			if (isInCSV(api, csvMatrix) is False and bk is False):
			#if (isInCSV(api, csvMatrix) is False):
				#print idx - 1, 'ffff', api
				continue

			if dictNum.has_key(api):
				dictNum[api] += 1
				hasApi = True
			else:
				dictNum[api] = 0
			
			response_time = re.search(ResponseTimePattern, line)
			res_time = response_time.group(1)
			res_time = float(res_time) 
			time_stamp = re.search(TimeStampPattern, line)
			time_st = time_stamp.group(1)[1:].split(' ')
			time_st = time_st[0] #String like 30/08/2016:10:30:00.000
			time_st = time_st.split('.') #Split string into to two parts like string 30/08/2016:10:30:00 and string 000
			time_st_1 = time.strptime(time_st[0], '%d/%m/%Y:%H:%M:%S') 
			#time_st_2 = float(time_st[1]) * 0.001 #Set time into the unit second
			time_st_1 = time.mktime(time_st_1)
			
			if time_st_1 < earlyTime:
				earlyTime = time_st_1
			if time_st_1 > latestTime:
				latestTime = time_st_1
			if hasApi is False:
				dict[api] = []
			dict[api].append((time_st_1, res_time))
			pass
			#print latestTime, earlyTime
			######
		except:
			#print idx, "format error"
			formaterror = True

	fp.close()
	cntRes = {} #Result of respose time of all the apis
	cntThr = {} #Result of latency of all the apis
	msize = 0 #Max size of data
	'''
		Calculation of the throughoutput and response_time
		by the delta time
	'''

	for key in dict:
		dict[key].sort(key=lambda x:x[0]) #Sort the data of all apis according to the timestamps 
		#print key, 'sort'
		startTime = int(earlyTime)
		endTime = int(startTime + delta)  
		cntRes[key] = []
		cntThr[key] = []
		i = 0
		l = len(dict[key])  
		cnt = 0
		ans_res = 0  
		#print startTime, latestTime
		pass
		while(i < l and startTime <= latestTime): 
			(time_st, res_time) = dict[key][i]
			pass
			if(startTime <= time_st and time_st <= endTime): 
				cnt += 1
				ans_res += res_time
				i += 1
				pass
			else:
				cntThr[key].append(cnt)
				if cnt is 0:
					cntRes[key].append(ans_res)
				else:
					cntRes[key].append(ans_res / (1.0 * cnt))
				pass	
				cnt = 0
				ans_res = 0
				startTime += delta
				endTime += delta
			pass
		cntThr[key].append(cnt)
		if cnt > 0:
			cntRes[key].append(ans_res / (1.0) * cnt)
		else :
			cntRes[key].append(ans_res)
		if msize < len(cntThr[key]):
			msize = len(cntThr[key])

	'''
		Fill up the sheet
	'''
	for key in cntThr:
		l = len(cntThr[key])
		d = msize - l
		while(d > 0):
			cntThr[key].append(0)
			d = d - 1
	for key in cntRes:
		l = len(cntRes[key])
		d = msize - l
		while(d > 0):
			cntRes[key].append(0.0)
			d = d - 1
			
	'''
		Write result to file named result.xls 
		Result of calculation of response time is stored as a sheet named res_time
		Result of calculation of throughoutput is stored as a sheet named throughoutput
	''' 
	print 'before wlwt'
	w = Workbook()
	ws = w.create_sheet('res_time')
	i = 1
	j = 1
	for key in cntRes:
		ws.cell(row=i, column=j).value = key
		j += 1
		for val in cntRes[key]:
			ws.cell(row=i, column=j).value = val
			j += 1
		i += 1
		j = 1
	ws2 = w.create_sheet('throughoutput')
	i = 1
	j = 1
	for key in cntThr:
		ws2.cell(row=i, column=j).value = key
		j += 1
		for val in cntThr[key]:
			ws2.cell(row=i, column=j).value = val
			j += 1
		i += 1
		j = 1
	resname = 'result_delta_'
	resname += str(delta)
	if(sysstr != "Windows"):
		fname = ""
	#w.save(fname + resname + '.xls')
	w.save(fname + resname + '.xlsx')
	print 'delta %d is done...\n' %(delta)
	
if __name__ == '__main__':
	InputFile = "new_access.log"
	#InputFile = "sample5000"
	delta = [1, 2, 3, 5, 10, 20]
	#delta = [2]
	csvFile = "grafana_data_export.csv"
	for val in delta:
		solve(val, InputFile, csvFile)
	print 'All done...'
	#solve(delta, InputFile, csvFile)

