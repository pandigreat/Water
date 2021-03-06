#-*-coding:utf-8-*-#
import platform
import sys
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl import load_workbook

'''
	Read the xlxs file
'''
def ReadFile(fileName):
	print "ReadFile"
	sysstr = platform.system()
	dir = sys.path[0]
	dir = dir.split('\\')
	preStr = ""
	for i in dir:
		preStr += i + "\\\\"
	if(sysstr == "Windows"):
		fileName = preStr + fileName
	print "fileName is ", fileName
	ncols = 0
	nrows = 0
	#try:
	wb = load_workbook(fileName) #Import the workbook
	sheetname = wb.get_sheet_names() #Get the names of all sheetssheetname = wb.get_sheet_names() #Get the names of all sheets
	sheet = wb.get_sheet_by_name(sheetname[0]) #Get the sheet
	ncols = sheet.max_column
	nrows = sheet.max_row
	print "norows:", nrows
	
	#except:
	#	print "Error For open workbook"
		
	
	Metrix = [[0 for col in range(ncols - 1)] for row in range(nrows)]
	KeyList = []
	
	for r in range(1, nrows + 1):
		for c in range(1, ncols + 1):
			val = sheet.cell(row=r, column=c).value
			if(c == 1):
				KeyList.append(val)
			else:
				Metrix[r - 1][c - 2] = float(val)
	
	return (KeyList, Metrix, nrows, ncols)
	
'''
	Record the result to a xlxs file 
'''
def RecordResult(KeyList, Correlation, P_values, nrows, ResultFileName):
	print "RecordResult"
	w = Workbook()
	ws = w.create_sheet('Correlation') #Insert a sheet from the head of the table
	for i in range(2, nrows + 2):
		for j in range(i, nrows + 2):
			ws.cell(row=i, column=j).value = Correlation[i - 2][j - 2]
			ws.cell(row=j, column=i).value = Correlation[j - 2][i - 2]
	for i in range(2, nrows + 2):
		val = KeyList[i - 2]
		ws.cell(row=1, column=i).value = val
		ws.cell(row=i, column=1).value = val
	
	ws = w.create_sheet('P_values')
	for i in range(2, nrows + 2):
		for j in range(i, nrows + 2):
			ws.cell(row=i, column=j).value = P_values[i - 2][j - 2]
			ws.cell(row=j, column=j).value = P_values[j - 2][i - 2]
	for i in range(2, nrows + 2):
		val = KeyList[i - 2]
		ws.cell(row=1, column=i).value = val
		ws.cell(row=i, column=1).value = val

	dir = sys.path[0]
	dir = dir.split('\\\\')
	preDir = ""
	for val in dir:
		preDir += val + '\\\\'
	sysstr = platform.system()
	if(sysstr == "Windows"):
		ResultFileName = preDir + ResultFileName
	
	w.save(ResultFileName)

'''
	Spearman Analysis for the data
'''
def SpearmanAnalysis(InputFileName, ResultFileName = "Result_Spearman.xlsx"):
	print "Spearman..." 
	
	KeyList, Metrix, nrows, ncols = ReadFile(InputFileName)
	Correlation = [[0 for col in range(nrows)] for row in range(nrows)]
	P_values = [[0 for col in range(nrows)] for row in range(nrows)]
	
	#print Correlation[0]
	
	for i in range(0, nrows):
		for j in range(i, nrows):
			cor, pval = stats.spearmanr(Metrix[i], Metrix[j])
			Correlation[i][j] = Correlation[j][i] = cor
			P_values[i][j] = P_values[j][i] = pval
	
	RecordResult(KeyList, Correlation, P_values, nrows, ResultFileName)
	print "Spearman Analysis has been done...\n\n"
	
def PearsonAnalysis(InputFileName, ResultFileName = "Result_Pearson.xlsx"):
	print "Pearson..."
	
	KeyList, Metrix, nrows, ncols = ReadFile(InputFileName)
	Correlation = [[0 for col in range(nrows)] for row in range(nrows)]
	P_values = [[0 for col in range(nrows)] for row in range(nrows)]
	
	for i in range(0, nrows):
		for j in range(0, nrows):
			cor, pval = stats.pearsonr(Metrix[i], Metrix[j])
			Correlation[i][j] = Correlation[j][i] = cor
			P_values[i][j] = P_values[j][i] = pval
	
	RecordResult(KeyList, Correlation, P_values, nrows, ResultFileName)
	print "Pearson Analysis has been done...\n\n"
	
if __name__=="__main__":
	
	print (platform.system())
	fileName = "circuit_break.xlsx"
	SpearmanAnalysis(fileName)
	PearsonAnalysis(fileName)
	
