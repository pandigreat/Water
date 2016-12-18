#-*-coding:utf-8-*-#
import sys
import platform
import numpy as np 
from decimal import *
from openpyxl import Workbook
from openpyxl import load_workbook
'''
	Read the xlxs file
'''
def ReadFile(fileName):
	print "ReadFile"
	sysstr = platform.system()
	dir = sys.path[0]
	dir = dir.split('\\')
	preStr = ""
	for i in dir:
		preStr += i + "\\\\"
	if(sysstr == "Windows"):
		fileName = preStr + fileName
	print "fileName is ", fileName
	ncols = 0
	nrows = 0
	#try:
	wb = load_workbook(fileName) #Import the workbook
	sheetname = wb.get_sheet_names() #Get the names of all sheetssheetname = wb.get_sheet_names() #Get the names of all sheets
	sheet = wb.get_sheet_by_name(sheetname[0]) #Get the sheet
	ncols = sheet.max_column
	nrows = sheet.max_row
	print "norows:", nrows
	
	Metrix = [[0 for col in range(ncols)] for row in range(nrows)]

	for r in range(1, nrows + 1):
		for c in range(1, ncols + 1):
			val = sheet.cell(row=r, column=c).value
			Metrix[r - 1][c - 1] = float(val)
	
	return (Metrix, nrows, ncols)
	
'''
	Record File
'''
def RecordResult(ResMetrix, nrows, rho):
	print "RecordResult"
	ResultFileName = "RecordResult" + "_rho_" + str(rho) + ".xlsx"
	w = Workbook()
	ws = w.create_sheet('Rho_Value_' + str(rho)) #Insert a sheet from the head of the table
	for i in range(0, nrows):
		for j in range(0, nrows):
			ws.cell(row=i+1, column=j+1).value = ResMetrix [i][j]

	dir = sys.path[0]
	dir = dir.split('\\\\')
	preDir = ""
	for val in dir:
		preDir += val + '\\\\'
	sysstr = platform.system()
	if(sysstr == "Windows"):
		ResultFileName = preDir + ResultFileName
	
	w.save(ResultFileName)

'''
	Ralation coefficient calcalation
'''
def RalationCoefficent(idx, Metrix, nrows, ncols, rho):
	print 'RalationCoefficent...', idx
	ResMetrix = [[0 for col in range(ncols)] for row in range(nrows)]
	li = []
	for i in range(nrows):
		if i is not idx:
			li.append(i)
	
	Min = Decimal(0x7fff)
	Max = Decimal(-0xffff)
	#print Metrix
	for i in li:
		for j in range(ncols):
			a = Decimal(Metrix[idx][j])
			b = Decimal(Metrix[i][j])
			#print float(a - b)
			val = abs(a- b)
			
			if val < Min:
				Min = val
			if val > Max:
				Max = val
	#print "Max Min", Max, Min
	for i in range(nrows):
		for j in range(ncols):
			if (i == idx):
				ResMetrix[i][j] = 1
			else:
				a = Decimal(Metrix[idx][j])
				b = Decimal(Metrix[i][j])
				val = abs(a - b)
				para1 = (Decimal(rho) * Decimal(Max) + Decimal(Min))
				para2 = (Decimal(rho) * Decimal(Max) + val)
				#if(para2  <= Decimal(0.0000000003)):
				#	para2 = para1
				#print "1, 2", para1, para2
				#print para1/para2
				ResMetrix[i][j] = para1 / para2
	
	Res = []
	
	for i in range(nrows):
		sum = 0
		for j in range(ncols):
			sum += ResMetrix[i][j]
		Res.append(sum / ncols)
	
	return Res
	
'''
	Gray Ralation Analysis
'''
def GrayRaltionAnalysis(InputFileName, rho, ResultFileName = "GrayRaltion_Result"):
	print 'GrayRaltionAnalysis...'
	(Metrix, nrows, ncols) = ReadFile(InputFileName)
	AverageMetrix = []
	for i in range(nrows):
		ave = 0.0
		for val in Metrix[i]:
			ave += val
		AverageMetrix.append(ave / ncols)
	
	for i in range(nrows):
		ave = AverageMetrix[i]
		for j in range(ncols):
			Metrix[i][j] = Metrix[i][j] / ave
	
	fp = open('f.txt', 'w')
	for i in range(nrows):
		for j in range(ncols):
			fp.write(str(Metrix[i][j]) + " ")
		fp.write("\n") 
	fp.close()
	
	RecordMetrix = []
	for i in range(nrows):
		res = RalationCoefficent(i, Metrix, nrows, ncols, rho)
		RecordMetrix.append(res)
	#print RecordMetrix
	RecordResult(RecordMetrix, nrows, rho)
	print 'Done...'
	
	
if __name__ == '__main__':
	InputFileName = "circuit_break_1.xlsx"
	rho = 0.5 #Normally 0.5, below 0.5463 precisely
	GrayRaltionAnalysis(InputFileName, rho)
	
	